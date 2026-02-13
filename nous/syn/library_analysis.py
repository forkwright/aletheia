#!/usr/bin/env python3

import pandas as pd
import openpyxl
from collections import Counter, defaultdict
import re
from datetime import datetime
import numpy as np

def analyze_physical_library():
    """Analyze Cody's physical book library from Excel file."""
    
    # Load the Excel file
    file_path = "/mnt/ssd/aletheia/theke/_reference/library/library_master.xlsx"
    
    print("Loading Excel file...")
    workbook = openpyxl.load_workbook(file_path)
    print(f"Sheets available: {workbook.sheetnames}")
    
    # Load the 'physical' sheet
    if 'physical' not in workbook.sheetnames:
        print("ERROR: 'physical' sheet not found!")
        return
    
    # Convert to pandas DataFrame
    df = pd.read_excel(file_path, sheet_name='physical')
    
    print(f"Loaded {len(df)} books from physical sheet")
    print(f"Columns: {list(df.columns)}")
    
    # Start analysis
    analysis = []
    analysis.append("# Cody's Physical Book Library Analysis")
    analysis.append(f"*Analysis generated on {datetime.now().strftime('%Y-%m-%d at %H:%M %Z')}*")
    analysis.append("")
    analysis.append(f"**Total Books**: {len(df)}")
    analysis.append("")
    
    # Show first few rows to understand structure
    analysis.append("## Data Structure Sample")
    analysis.append("```")
    analysis.append(df.head(3).to_string())
    analysis.append("```")
    analysis.append("")
    
    # 1. Genre breakdown
    analysis.append("## 1. Genre Breakdown")
    if 'Genre' in df.columns:
        genres = df['Genre'].dropna()
        genre_counts = genres.value_counts()
        total_with_genre = len(genres)
        
        analysis.append(f"**Books with genre information**: {total_with_genre} of {len(df)} ({total_with_genre/len(df)*100:.1f}%)")
        analysis.append("")
        analysis.append("| Genre | Count | Percentage |")
        analysis.append("|-------|-------|------------|")
        
        for genre, count in genre_counts.head(20).items():
            pct = count / total_with_genre * 100
            analysis.append(f"| {genre} | {count} | {pct:.1f}% |")
        
        if len(genre_counts) > 20:
            others = genre_counts[20:].sum()
            pct = others / total_with_genre * 100
            analysis.append(f"| Other genres ({len(genre_counts) - 20}) | {others} | {pct:.1f}% |")
    else:
        analysis.append("*No Genre column found*")
    
    analysis.append("")
    
    # 2. Read Status breakdown
    analysis.append("## 2. Read Status Breakdown")
    if 'ReadStatus' in df.columns:
        read_status = df['ReadStatus']
        read_count = (read_status == 1).sum()
        unread_count = (read_status == 0).sum()
        null_count = read_status.isna().sum()
        other_count = len(df) - read_count - unread_count - null_count
        
        analysis.append(f"- **Read** (ReadStatus = 1): {read_count} books ({read_count/len(df)*100:.1f}%)")
        analysis.append(f"- **Unread** (ReadStatus = 0): {unread_count} books ({unread_count/len(df)*100:.1f}%)")
        if null_count > 0:
            analysis.append(f"- **No status** (ReadStatus = null): {null_count} books ({null_count/len(df)*100:.1f}%)")
        if other_count > 0:
            analysis.append(f"- **Other status**: {other_count} books ({other_count/len(df)*100:.1f}%)")
        
        if read_count + unread_count > 0:
            analysis.append(f"- **Reading completion rate**: {read_count/(read_count+unread_count)*100:.1f}%")
        else:
            analysis.append(f"- **Reading completion rate**: Cannot calculate (no clear read/unread status)")
    else:
        analysis.append("*No ReadStatus column found*")
    
    analysis.append("")
    
    # 3. Top 20 authors by book count
    analysis.append("## 3. Top 20 Authors by Book Count")
    
    # Check for Author column or Author First/Last columns
    if 'Author' in df.columns:
        authors = df['Author'].dropna()
        author_counts = authors.value_counts()
    elif 'Author First' in df.columns and 'Author Last' in df.columns:
        # Combine first and last name
        df_authors = df[['Author First', 'Author Last']].copy()
        df_authors['Full Author'] = df_authors['Author First'].fillna('') + ' ' + df_authors['Author Last'].fillna('')
        df_authors['Full Author'] = df_authors['Full Author'].str.strip()
        authors = df_authors['Full Author'][df_authors['Full Author'] != '']
        author_counts = authors.value_counts()
    else:
        authors = None
        
    if authors is not None and len(author_counts) > 0:
        analysis.append("| Rank | Author | Books |")
        analysis.append("|------|--------|-------|")
        
        for i, (author, count) in enumerate(author_counts.head(20).items(), 1):
            analysis.append(f"| {i} | {author} | {count} |")
        
        analysis.append("")
        analysis.append(f"**Total unique authors**: {len(author_counts)}")
        analysis.append(f"**Average books per author**: {len(authors) / len(author_counts):.1f}")
    else:
        analysis.append("*No Author information found*")
    
    analysis.append("")
    
    # 4. Series inventory
    analysis.append("## 4. Series Inventory")
    if 'Series' in df.columns:
        series_data = df['Series'].dropna()
        series_counts = series_data.value_counts()
        
        analysis.append(f"**Books in series**: {len(series_data)} of {len(df)} ({len(series_data)/len(df)*100:.1f}%)")
        analysis.append(f"**Unique series**: {len(series_counts)}")
        analysis.append("")
        analysis.append("### Top Series by Volume Count")
        analysis.append("| Series | Volumes |")
        analysis.append("|--------|---------|")
        
        for series, count in series_counts.head(15).items():
            analysis.append(f"| {series} | {count} |")
    else:
        analysis.append("*No Series column found*")
    
    analysis.append("")
    
    # 5. Publication date distribution
    analysis.append("## 5. Publication Date Distribution")
    pub_date_cols = [col for col in df.columns if 'date' in col.lower() or 'year' in col.lower() or 'published' in col.lower()]
    
    if pub_date_cols:
        analysis.append(f"**Found potential date columns**: {pub_date_cols}")
        
        # Try to extract years from the first date column
        date_col = pub_date_cols[0]
        dates = df[date_col].dropna()
        
        # Extract years
        years = []
        for date in dates:
            if pd.isna(date):
                continue
            try:
                if isinstance(date, (int, float)):
                    if 1800 <= date <= 2030:
                        years.append(int(date))
                elif isinstance(date, str):
                    # Try to extract 4-digit year
                    year_match = re.search(r'\b(19|20)\d{2}\b', str(date))
                    if year_match:
                        years.append(int(year_match.group()))
                else:
                    # Handle datetime objects
                    years.append(date.year if hasattr(date, 'year') else None)
            except:
                continue
        
        if years:
            years = [y for y in years if y is not None and 1800 <= y <= 2030]
            
            # Group by decade
            decades = {}
            for year in years:
                decade = (year // 10) * 10
                decades[decade] = decades.get(decade, 0) + 1
            
            analysis.append(f"**Books with publication dates**: {len(years)} of {len(df)}")
            analysis.append("")
            analysis.append("### By Decade")
            analysis.append("| Decade | Count | Percentage |")
            analysis.append("|--------|-------|------------|")
            
            for decade in sorted(decades.keys()):
                count = decades[decade]
                pct = count / len(years) * 100
                analysis.append(f"| {decade}s | {count} | {pct:.1f}% |")
        else:
            analysis.append("*Could not parse publication dates*")
    else:
        analysis.append("*No publication date columns found*")
    
    analysis.append("")
    
    # 6. Notable patterns and observations
    analysis.append("## 6. Notable Patterns and Observations")
    
    observations = []
    
    # Collection size analysis
    observations.append(f"- **Collection size**: {len(df)} books is substantial for a 12x10 office space (1.5 walls)")
    
    # Density calculation
    books_per_wall = len(df) / 1.5
    observations.append(f"- **Storage density**: ~{books_per_wall:.0f} books per wall section")
    
    # Genre analysis
    if 'Genre' in df.columns:
        top_genre = df['Genre'].value_counts().index[0] if not df['Genre'].isna().all() else "Unknown"
        top_genre_count = df['Genre'].value_counts().iloc[0] if not df['Genre'].isna().all() else 0
        observations.append(f"- **Dominant genre**: {top_genre} ({top_genre_count} books)")
    
    # Author concentration
    if 'Author' in df.columns:
        author_counts = df['Author'].value_counts()
        if len(author_counts) > 0:
            top_author = author_counts.index[0]
            top_author_count = author_counts.iloc[0]
            observations.append(f"- **Most collected author**: {top_author} ({top_author_count} books)")
            
            # Calculate author concentration
            top_10_authors = author_counts.head(10).sum()
            pct_top_10 = top_10_authors / len(df) * 100
            observations.append(f"- **Author concentration**: Top 10 authors represent {pct_top_10:.1f}% of collection")
    
    # Series vs standalone
    if 'Series' in df.columns:
        series_books = df['Series'].notna().sum()
        standalone_books = len(df) - series_books
        observations.append(f"- **Series vs Standalone**: {series_books} books in series, {standalone_books} standalone ({series_books/len(df)*100:.1f}% series)")
    
    # Reading progress
    if 'ReadStatus' in df.columns:
        read_count = (df['ReadStatus'] == 1).sum()
        unread_count = (df['ReadStatus'] == 0).sum()
        null_count = df['ReadStatus'].isna().sum()
        if read_count + unread_count > 0:
            read_rate = read_count / (read_count + unread_count) * 100
            observations.append(f"- **Reading progress**: {read_rate:.1f}% completion rate ({read_count} read, {unread_count} unread)")
            
            # Estimate reading backlog
            if unread_count > 0:
                # Assume 1-2 books per month reading rate
                months_backlog = unread_count / 1.5
                years_backlog = months_backlog / 12
                observations.append(f"- **Reading backlog**: ~{years_backlog:.1f} years at 1.5 books/month")
        else:
            observations.append(f"- **Reading status**: {read_count} marked as read, {null_count} books without status")
    
    # Space optimization observations
    total_books = len(df)
    if total_books > 500:
        observations.append(f"- **Space utilization**: With {total_books} books in limited wall space, likely using floor-to-ceiling shelving")
    
    # Collection maturity
    if pub_date_cols and years:
        avg_year = sum(years) / len(years)
        current_year = datetime.now().year
        avg_age = current_year - avg_year
        observations.append(f"- **Collection age**: Average publication year ~{avg_year:.0f} (books are ~{avg_age:.0f} years old on average)")
    
    for obs in observations:
        analysis.append(obs)
    
    analysis.append("")
    
    # Summary statistics
    analysis.append("## Summary Statistics")
    analysis.append(f"- **Total books**: {len(df)}")
    if 'Genre' in df.columns:
        genre_count = df['Genre'].nunique()
        analysis.append(f"- **Genres represented**: {genre_count}")
    if 'Author' in df.columns:
        author_count = df['Author'].nunique()
        analysis.append(f"- **Unique authors**: {author_count}")
    if 'Series' in df.columns:
        series_count = df['Series'].nunique()
        analysis.append(f"- **Series**: {series_count}")
    
    analysis.append("")
    analysis.append("---")
    analysis.append("*End of Analysis*")
    
    return '\n'.join(analysis), df

if __name__ == "__main__":
    try:
        analysis_text, df = analyze_physical_library()
        print("Analysis completed successfully!")
        print(f"Analyzed {len(df)} books")
        
        # Write to file
        output_file = "/mnt/ssd/aletheia/theke/_reference/library/physical_library_analysis.md"
        with open(output_file, 'w', encoding='utf-8') as f:
            f.write(analysis_text)
        print(f"Analysis written to: {output_file}")
        
        # Print summary
        print("\n" + "="*50)
        print("QUICK SUMMARY")
        print("="*50)
        
        total_books = len(df)
        print(f"Total books: {total_books}")
        
        if 'ReadStatus' in df.columns:
            read_count = (df['ReadStatus'] == 1).sum()
            unread_count = (df['ReadStatus'] == 0).sum()
            null_count = df['ReadStatus'].isna().sum()
            read_rate = read_count / (read_count + unread_count) * 100 if (read_count + unread_count) > 0 else 0
            print(f"Read: {read_count}, Unread: {unread_count}, No status: {null_count} ({read_rate:.1f}% read)")
        
        if 'Genre' in df.columns:
            top_genre = df['Genre'].value_counts().index[0] if not df['Genre'].isna().all() else "Unknown"
            top_genre_count = df['Genre'].value_counts().iloc[0] if not df['Genre'].isna().all() else 0
            print(f"Top genre: {top_genre} ({top_genre_count} books)")
        
        if 'Author' in df.columns:
            top_author = df['Author'].value_counts().index[0] if not df['Author'].isna().all() else "Unknown"
            top_author_count = df['Author'].value_counts().iloc[0] if not df['Author'].isna().all() else 0
            print(f"Top author: {top_author} ({top_author_count} books)")
            
    except Exception as e:
        print(f"Error during analysis: {e}")
        import traceback
        traceback.print_exc()